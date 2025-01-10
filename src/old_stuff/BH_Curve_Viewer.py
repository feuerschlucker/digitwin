import numpy as np
from PySide6.QtWidgets import QApplication, QMainWindow, QVBoxLayout, QHBoxLayout, QPushButton, QWidget
import pyqtgraph as pg
from scipy.signal import savgol_filter
from scipy.integrate import trapezoid
import copy
import openpyxl


# ------------------------------------------------ calculations --------------------------------------------------------

def select_periods(data, periods):
    periods_ = []
    for period in periods:
        periods_.append(data['periods'][period])
    data['periods'] = periods_


def drift_correction(data):
    # drift correction of flux
    start_idx = data['periods'][0][0]
    end_idx = data['periods'][-1][1]
    slope = (data['Φ2 [mWb]'][end_idx] - data['Φ2 [mWb]'][start_idx]) / (end_idx - start_idx)
    correction = (np.arange(len(data['Φ2 [mWb]'])) - len(data['Φ2 [mWb]']) / 2) * slope
    data['Φ2 [mWb]'] = data['Φ2 [mWb]'] - correction


def filter_data(data, filter_type):
    if filter_type == 'svg':
        window = 200
        data['Φ1 [mWb]'] = savgol_filter(data['Φ1 [mWb]'], window, 2)
        data['Φ2 [mWb]'] = savgol_filter(data['Φ2 [mWb]'], window, 2)
        data['I_prim [A]'] = savgol_filter(data['I_prim [A]'], window, 2)


def average_periods(data):
    max_len = 0
    Φ1_slices = []
    Φ2_slices = []
    I_prim_slices = []
    for period in data['periods']:
        Φ1_slices.append(data['Φ1 [mWb]'][period[0]:period[1]])
        Φ2_slices.append(data['Φ2 [mWb]'][period[0]:period[1]])
        I_prim_slices.append(data['I_prim [A]'][period[0]:period[1]])
        if len(data['Φ1 [mWb]'][period[0]:period[1]]) > max_len:
            max_len = len(data['Φ1 [mWb]'][period[0]:period[1]])

    def interp1d(array: np.ndarray, new_len: int) -> np.ndarray:
        la = len(array)
        return np.interp(np.linspace(0, la - 1, num=new_len), np.arange(la), array)

    # average
    for i in range(len(I_prim_slices)):
        if i == 0:
            Φ1_sum = interp1d(Φ1_slices[i], new_len=max_len)
            Φ2_sum = interp1d(Φ2_slices[i], new_len=max_len)
            I_prim_sum = interp1d(I_prim_slices[i], new_len=max_len)
        else:
            Φ1_sum = Φ1_sum + interp1d(Φ1_slices[i], new_len=max_len)
            Φ2_sum = Φ2_sum + interp1d(Φ2_slices[i], new_len=max_len)
            I_prim_sum = I_prim_sum + interp1d(I_prim_slices[i], new_len=max_len)
    data['Φ1 [mWb]'] = Φ1_sum / len(I_prim_slices)
    data['Φ2 [mWb]'] = Φ2_sum / len(I_prim_slices)
    data['I_prim [A]'] = I_prim_sum / len(I_prim_slices)
    # adjust period data
    data['periods'] = [(0, max_len)]


def center_curve(data):
    # todo is this allowed?
    # center vertical
    for period in data['periods']:
        p_data_1 = data['Φ1 [mWb]'][period[0]:period[1]]
        p_data_2 = data['Φ2 [mWb]'][period[0]:period[1]]
        min_phi1 = np.min(p_data_1)
        max_phi1 = np.max(p_data_1)
        min_phi2 = np.min(p_data_2)
        max_phi2 = np.max(p_data_2)
        #
        offset_1 = max_phi1 - ((max_phi1 - min_phi1) / 2)
        offset_2 = max_phi2 - ((max_phi2 - min_phi2) / 2)
        #
        data['Φ1 [mWb]'][period[0]:period[1]] = p_data_1 - offset_1
        data['Φ2 [mWb]'][period[0]:period[1]] = p_data_2 - offset_2


def calc_curve_params(data):
    for period in data['periods']:
        theta = data['theta'][period[0]:period[1]]
        phi = data['Φ2 [mWb]'][period[0]:period[1]]
        # get upper curve and lower curve
        peak_ind_1 = np.argmax(theta)
        peak_ind_2 = np.argmin(theta)
        # up
        upper_curve_ind = np.array(range(peak_ind_1, peak_ind_2), dtype=int)
        theta_up = theta[upper_curve_ind]
        phi_up = phi[upper_curve_ind]
        # low
        lower_curve_ind1 = np.array(range(peak_ind_2, len(theta)), dtype=int)
        lower_curve_ind2 = np.array(range(0, peak_ind_1), dtype=int)
        lower_curve_ind = np.concatenate((lower_curve_ind1, lower_curve_ind2), axis=0)
        theta_lo = theta[lower_curve_ind]
        phi_lo = phi[lower_curve_ind]
        # # calc area
        area_1 = np.abs(trapezoid(phi_up + 2, x=theta_up))
        area_2 = trapezoid(phi_lo + 2, x=theta_lo)
        area = area_1 - area_2
        # calc Saturation
        saturation = np.max(phi)
        # calc coercitivity
        zero_index = np.argmax(phi_up <= 0)
        coercitivity = theta_up[zero_index]
        # calc remanence
        zero_index = np.argmax(theta_up <= 0)
        remanence = phi_up[zero_index]
        # calc coercitivity 2
        zero_index = np.argmax(phi_lo >= 0)
        coercitivity_2 = theta_lo[zero_index]
        # width
        curve_width = coercitivity_2 - coercitivity
        # print data
        print(f'area = {area} [mWbA], '
              f'saturation = {saturation} [mWb], '
              f'coercitivity = {coercitivity} [A], '
              f'coercitivity_2 = {coercitivity_2} [A], '
              f'remanence = {remanence} [mWbA], '
              f'curve_width = {curve_width} [A]')
        # store in dict
        curve_parameters = {}
        curve_parameters['area [mWbA]'] = area
        curve_parameters['saturation [mWb]'] = saturation
        curve_parameters['coercitivity [A]'] = coercitivity
        curve_parameters['coercitivity_2 [A]'] = coercitivity_2
        curve_parameters['remanence [mWb]'] = remanence
        curve_parameters['curve_width [A]'] = curve_width
        # # add data to excel
        # file_path = 'parameter_comparison.xlsx'
        # wb = openpyxl.load_workbook(file_path)
        # sheet = wb.active
        # # Find the next empty row
        # next_row = sheet.max_row + 1
        # # Add data to the next empty row
        # mp = data['measurement_protocol']
        # col_data = [area, saturation, coercitivity, coercitivity_2, remanence, curve_width, mp['Meas_ID'],
        #             mp['Waveform'], mp['Amplitude'], mp['Yoke'], mp['Sample_ID'], mp['Frequency'],
        #             mp['Duration'], mp['Paste']]
        # for col, value in enumerate(col_data, start=1):
        #     sheet.cell(row=next_row, column=col, value=value)
        # # Save the workbook
        # wb.save(file_path)
        return curve_parameters


def add_to_comparison(self):
    data = {}
    for ref in self.plot_references.keys():
        if not ('line' in ref):
            data[ref] = {}
            data[ref]['x_values'] = self.plot_references[ref].xData
            data[ref]['y_values'] = self.plot_references[ref].yData
            data[ref]['linewidth'] = self.plot_references[ref].opts['pen'].width()
            data[ref]['Meas_ID'] = self.data['Meas_ID']
    self.comparisonWindow.addData(data)
    self.comparisonWindow.update()
    self.comparisonWindow.show()


# ------------------------------------------------- make plot ----------------------------------------------------------

def plot_BH_curve(self):
    data = copy.deepcopy(self.data)
    # clear old data
    for ref in self.plot_references.keys():
        if ref == 'zero_line_h' or ref == 'zero_line_v':
            continue
        else:
            self.widgets_dict['plot_area'].removeItem(self.plot_references[ref])

    if self.plot_config['used_periods']:
        select_periods(data, self.plot_config['used_periods'])

    if self.plot_config['drift_correction']:
        drift_correction(data)

    if self.plot_config['filter'] is not None:
        filter_data(data, self.plot_config['filter'])

    if self.plot_config['average_periods']:
        average_periods(data)

    if self.plot_config['center']:
        center_curve(data)

    # calculate theta (Durchflutung)
    data['theta'] = 1000 * (data['I_prim [A]'] - data['I_offset'])  # 1000 coil turns on sec
    #
    colors = ['#ff5252', '#e8971c', '#faf861', '#41cc99', '#4ceefc', '#4c78fc', '#3832b3', '#321363', '#f21882']
    for i, p in enumerate(data['periods']):
        self.plot_references[f'period {i}'] = self.widgets_dict['plot_area'].plot(data['theta'][p[0]: p[1]],
                                                                                  data['Φ2 [mWb]'][p[0]: p[1]],
                                                                                  pen=pg.mkPen(colors[i], width=1))
    if self.plot_config['plot_inverted']:
        for i, p in enumerate(data['periods']):
            self.plot_references[f'period {i} inverted'] = self.widgets_dict['plot_area'].plot(-data['theta'][p[0]: p[1]],
                                                                                      -data['Φ2 [mWb]'][p[0]: p[1]],
                                                                                      pen=pg.mkPen(colors[i], width=1))

    if self.plot_config['add_to_comparison']:
        add_to_comparison(self)

    # calculate curve parameters like area, remanence, coercivity, ...
    return calc_curve_params(data)


# ---------------------------------------------- define buttons and GUI ------------------------------------------------

def get_comparison_button(self):
    # Database loading UI
    main_layout = QVBoxLayout()
    self.widgets_dict['comparison_button'] = QPushButton('Add to Comparison')
    self.widgets_dict['comparison_button'].setStyleSheet("""QPushButton {background-color: #c1dae6;
                                                                        border-radius: 4px;
                                                                        max-width: 150px;
                                                                        height: 60px;}""")
    self.widgets_dict['comparison_button'].clicked.connect(lambda: add_to_comparison(self))
    main_layout.addWidget(self.widgets_dict['comparison_button'])
    return main_layout


def get_average_button(self):
    # Database loading UI
    main_layout = QVBoxLayout()
    self.widgets_dict['average_data_button'] = QPushButton('Average Periods')
    self.widgets_dict['average_data_button'].setStyleSheet("""QPushButton {background-color: #c1dae6;
                                                                        border-radius: 4px;
                                                                        max-width: 150px;
                                                                        height: 60px;}""")
    self.widgets_dict['average_data_button'].clicked.connect(lambda: average_periods_(self))
    main_layout.addWidget(self.widgets_dict['average_data_button'])
    return main_layout


def create_task_bar(self):
    layout = QHBoxLayout()
    layout.addLayout(get_average_button(self))
    layout.addLayout(get_comparison_button(self))
    return layout


def get_default_plot_config():
    dct = {}
    dct['used_periods'] = [1]  # [] = all periods are used, (enumeration starts with 0)
    dct['drift_correction'] = True
    dct['filter'] = 'svg'  # filters: 'svg' (savgol filter), None = no filter
    dct['average_periods'] = False
    dct['add_to_comparison'] = False
    dct['center'] = True
    dct['plot_inverted'] = False
    dct['save_option'] = None  # None, png, svg, matplotlib, ...
    return dct


class BHCurveWindow(QMainWindow):
    def __init__(self, comparisonWindow):
        super().__init__()
        # window title
        self.setWindowTitle('BH-Curve-Viewer')

        # comparison window object
        self.comparisonWindow = comparisonWindow

        # make plot options dict
        self.plot_config = get_default_plot_config()

        # setup widgets dict
        self.widgets_dict = {}

        # plots
        self.plot_references = {}

        # Set up the main layout
        self.widgets_dict['plot_area'] = pg.PlotWidget()  # Create a plot widget
        self.widgets_dict['plot_area'].setBackground('black')
        # Set lines to cross at zero
        self.plot_references['zero_line_h'] = (pg.InfiniteLine(pos=0, angle=0, pen=pg.mkPen('white', width=1)))
        self.plot_references['zero_line_v'] = (pg.InfiniteLine(pos=0, angle=90, pen=pg.mkPen('white', width=1)))
        self.widgets_dict['plot_area'].addItem(self.plot_references['zero_line_h'])
        self.widgets_dict['plot_area'].addItem(self.plot_references['zero_line_v'])
        # add grid
        self.widgets_dict['plot_area'].showGrid(x=True, y=True, alpha=0.3)
        # Add axis labels
        self.widgets_dict['plot_area'].setLabel('left', 'Φ [mWb]', color='white', size='12pt')
        self.widgets_dict['plot_area'].setLabel('bottom', 'Θ [A]', color='white', size='12pt')

        #
        main_layout = QVBoxLayout()
        main_layout.addLayout(create_task_bar(self))
        main_layout.addWidget(self.widgets_dict['plot_area'])

        # Set the central widget
        central_widget = QWidget()
        central_widget.setLayout(main_layout)
        self.setCentralWidget(central_widget)

        # data added to window
        self.data = None  # Store plot data

    def addData(self, data):
        self.data = data
        curve_parameters = plot_BH_curve(self)
        return curve_parameters
